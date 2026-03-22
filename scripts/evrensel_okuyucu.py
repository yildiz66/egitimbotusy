"""
Evrensel Dosya Okuyucu
======================
Kullanıcının yüklediği her formattan veri çeker.

Desteklenen formatlar:
  Ders programı : .xlsx, .xls, .docx, .pdf, .png, .jpg, .jpeg
  Yıllık plan   : .xlsx, .xls, .docx, .pdf
  Önceki evrak  : .docx, .pdf

Görsel formatlar için OCR (Tesseract) kullanılır.
Tesseract kurulu değilse PIL ile temel metin çıkarımı denenir.
"""

import re
import json
from pathlib import Path

# ── Kütüphane kontrolleri ──
try:
    from docx import Document as DocxDoc
    DOCX_OK = True
except ImportError:
    DOCX_OK = False

try:
    import openpyxl
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

try:
    import pdfplumber
    PDF_OK = True
except ImportError:
    PDF_OK = False

try:
    from PIL import Image
    PIL_OK = True
except ImportError:
    PIL_OK = False

try:
    import pytesseract
    # Tesseract kurulu mu test et
    pytesseract.get_tesseract_version()
    OCR_OK = True
except Exception:
    OCR_OK = False


# ══════════════════════════════════════════════════════
# DÜŞÜK SEVİYE: ham metin çıkarıcılar
# ══════════════════════════════════════════════════════

def word_oku(dosya: Path) -> str:
    if not DOCX_OK:
        return ""
    doc = DocxDoc(dosya)
    satirlar = []
    for p in doc.paragraphs:
        if p.text.strip():
            satirlar.append(p.text.strip())
    # Tablolar da oku
    for tablo in doc.tables:
        for satir in tablo.rows:
            hucre_metinleri = [h.text.strip() for h in satir.cells if h.text.strip()]
            if hucre_metinleri:
                satirlar.append(" | ".join(hucre_metinleri))
    return "\n".join(satirlar)


def excel_oku(dosya: Path) -> str:
    """Excel'den tüm sayfaları metin olarak döndürür."""
    if not EXCEL_OK:
        return ""
    wb = openpyxl.load_workbook(dosya, data_only=True)
    satirlar = []
    for sayfa_adi in wb.sheetnames:
        ws = wb[sayfa_adi]
        satirlar.append(f"[SAYFA: {sayfa_adi}]")
        for satir in ws.iter_rows(values_only=True):
            hucre = [str(h).strip() for h in satir if h is not None and str(h).strip()]
            if hucre:
                satirlar.append(" | ".join(hucre))
    return "\n".join(satirlar)


def excel_tablo_oku(dosya: Path) -> list[list]:
    """Excel'den 2D liste olarak döndürür (program çözümleyici için)."""
    if not EXCEL_OK:
        return []
    wb = openpyxl.load_workbook(dosya, data_only=True)
    ws = wb.active
    satirlar = []
    for satir in ws.iter_rows(values_only=True):
        if any(h is not None for h in satir):
            satirlar.append([str(h).strip() if h is not None else "" for h in satir])
    return satirlar


def pdf_oku(dosya: Path) -> str:
    if not PDF_OK:
        return ""
    metin = ""
    with pdfplumber.open(dosya) as pdf:
        for sayfa in pdf.pages:
            t = sayfa.extract_text()
            if t:
                metin += t + "\n"
    return metin.strip()


def gorsel_oku(dosya: Path) -> str:
    """Görsel dosyadan OCR ile metin çıkarır."""
    if not PIL_OK:
        return ""
    img = Image.open(dosya)
    if OCR_OK:
        # Türkçe + İngilizce OCR
        try:
            metin = pytesseract.image_to_string(img, lang="tur+eng")
            return metin.strip()
        except Exception:
            pass
    # OCR yoksa boş döndür, uyarı ver
    print(f"   ⚠  OCR bulunamadı — {dosya.name} görsel dosyası metin olarak okunamadı.")
    print("      Tesseract kurmak için: sudo apt-get install tesseract-ocr tesseract-ocr-tur")
    return ""


def dosya_oku(dosya: Path) -> str:
    """Uzantıya göre doğru okuyucuyu seçer."""
    ext = dosya.suffix.lower()
    if ext == ".docx":
        return word_oku(dosya)
    elif ext in (".xlsx", ".xls"):
        return excel_oku(dosya)
    elif ext == ".pdf":
        return pdf_oku(dosya)
    elif ext in (".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".webp"):
        return gorsel_oku(dosya)
    return ""


def klasor_tara(klasor: Path) -> list[Path]:
    """Klasördeki tüm desteklenen dosyaları listeler."""
    uzantilar = {".docx", ".xlsx", ".xls", ".pdf", ".png", ".jpg", ".jpeg"}
    return sorted([f for f in klasor.rglob("*") if f.suffix.lower() in uzantilar])


# ══════════════════════════════════════════════════════
# DERS PROGRAMI ÇÖZÜMLEYICI
# Word / Excel / PDF / Görsel → {"Pazartesi": [...], ...}
# ══════════════════════════════════════════════════════

GUNLER = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]
GUN_KISALTMA = {
    "pzt": "Pazartesi", "sal": "Salı", "çar": "Çarşamba",
    "per": "Perşembe", "cum": "Cuma",
    "mon": "Pazartesi", "tue": "Salı", "wed": "Çarşamba",
    "thu": "Perşembe",  "fri": "Cuma",
}

def _gun_bul(metin: str) -> str | None:
    m = metin.lower().strip()
    for k, v in GUN_KISALTMA.items():
        if k in m:
            return v
    for g in GUNLER:
        if g.lower() in m:
            return g
    return None

def _sinif_bul(metin: str) -> str | None:
    """6-C, 7A, 8/B gibi formatları yakalar."""
    m = re.search(r"\b([5-8])\s*[-/]?\s*([A-Ea-e])\b", metin)
    return f"{m.group(1)}-{m.group(2).upper()}" if m else None

def _saat_bul(metin: str) -> str | None:
    m = re.search(r"\b([0-9]{1,2})[:\.]([0-9]{2})\b", metin)
    return f"{int(m.group(1)):02d}:{m.group(2)}" if m else None


def excel_program_coz(dosya: Path) -> dict:
    """
    Excel ders programı çözümleyici.
    Beklenen format:
      Satır 1: başlık (Saat | Pazartesi | Salı | ...)
      Satır 2+: 08:00 | 7-A | 6-C | ...
    Alternatif: dikey format da denenir.
    """
    satirlar = excel_tablo_oku(dosya)
    if not satirlar:
        return {}

    program = {g: [] for g in GUNLER}
    baslik_satiri = satirlar[0]

    # Başlık satırında gün adı ara
    gun_sutunlari = {}   # sütun_idx → gün_adı
    for ci, h in enumerate(baslik_satiri):
        gun = _gun_bul(h)
        if gun:
            gun_sutunlari[ci] = gun

    if gun_sutunlari:
        # Yatay format: her satır = bir saat dilimi
        for satir in satirlar[1:]:
            if not satir:
                continue
            saat = _saat_bul(satir[0]) if satir else None
            for ci, gun in gun_sutunlari.items():
                if ci < len(satir) and satir[ci].strip():
                    sinif = _sinif_bul(satir[ci])
                    if sinif:
                        program[gun].append({
                            "saat":  saat or satir[0],
                            "sinif": sinif,
                        })
    else:
        # Dikey format veya serbest yerleşim: satır satır tüm metni tara
        for satir in satirlar:
            metin = " ".join(satir)
            gun   = _gun_bul(metin)
            sinif = _sinif_bul(metin)
            saat  = _saat_bul(metin)
            if gun and sinif:
                program[gun].append({
                    "saat":  saat or "08:00",
                    "sinif": sinif,
                })

    # Boş günleri temizle
    return {g: v for g, v in program.items() if v}


def metin_program_coz(metin: str) -> dict:
    """
    Düz metinden (Word/PDF/OCR) ders programı çıkarır.
    Satır bazlı + bağlam bazlı ayrıştırma.
    """
    program = {g: [] for g in GUNLER}
    aktif_gun = None
    aktif_saat = None

    for satir in metin.split("\n"):
        satir = satir.strip()
        if not satir:
            continue

        gun = _gun_bul(satir)
        if gun:
            aktif_gun = gun

        saat = _saat_bul(satir)
        if saat:
            aktif_saat = saat

        sinif = _sinif_bul(satir)
        if sinif and aktif_gun:
            program[aktif_gun].append({
                "saat":  aktif_saat or "08:00",
                "sinif": sinif,
            })

    return {g: v for g, v in program.items() if v}


def program_oku(klasor: Path) -> dict:
    """
    Klasördeki ilk uygun dosyadan ders programını okur.
    Öncelik: JSON > Excel > Word > PDF > Görsel
    """
    # 1. JSON (elle girilmiş)
    j = klasor / "program.json"
    if j.exists():
        with open(j, encoding="utf-8") as f:
            return json.load(f)

    # 2. Excel (en güvenilir format)
    for f in list(klasor.glob("*.xlsx")) + list(klasor.glob("*.xls")):
        sonuc = excel_program_coz(f)
        if sonuc:
            print(f"   ✓  Ders programı Excel'den okundu: {f.name}")
            return sonuc

    # 3. Word / PDF / Görsel — metin bazlı çözümleme
    for f in klasor_tara(klasor):
        metin = dosya_oku(f)
        if metin:
            sonuc = metin_program_coz(metin)
            if sonuc:
                print(f"   ✓  Ders programı {f.suffix.upper()} dosyasından okundu: {f.name}")
                return sonuc

    print("   ⚠  Ders programı dosyası bulunamadı → örnek program kullanılıyor.")
    return {}


# ══════════════════════════════════════════════════════
# YILLIK PLAN ÇÖZÜMLEYICI
# Word / Excel / PDF → {hafta: {konu, kazanim, unite}}
# ══════════════════════════════════════════════════════

def yillik_plan_oku(klasor: Path, model: str = "normal") -> dict:
    """
    Yıllık planı okur ve hafta → {konu, kazanim, unite} sözlüğü döndürür.
    model = 'maarif' ise alt klasörü de arar.
    """
    aranacak = [klasor]
    if model == "maarif":
        aranacak.insert(0, klasor / "maarif")

    for dizin in aranacak:
        if not dizin.exists():
            continue
        for f in klasor_tara(dizin):
            if f.suffix.lower() in (".xlsx", ".xls"):
                sonuc = _excel_yillik_coz(f)
            else:
                metin = dosya_oku(f)
                sonuc = _metin_yillik_coz(metin) if metin else {}
            if sonuc:
                print(f"   ✓  Yıllık plan okundu ({model}): {f.name}")
                return sonuc

    return {}


def _excel_yillik_coz(dosya: Path) -> dict:
    """Excel yıllık planından hafta/konu/kazanım tablosunu çeker."""
    satirlar = excel_tablo_oku(dosya)
    plan = {}
    hafta_col = konu_col = kazanim_col = unite_col = None

    for satir in satirlar[:3]:   # Başlık satırını bul
        for ci, h in enumerate(satir):
            hl = h.lower()
            if "hafta" in hl:       hafta_col   = ci
            if "konu" in hl:        konu_col    = ci
            if "kazanım" in hl or "kazanim" in hl: kazanim_col = ci
            if "ünite" in hl or "unite" in hl:     unite_col   = ci
        if hafta_col is not None:
            break

    if hafta_col is None:
        return {}

    aktif_hafta = None
    for satir in satirlar[1:]:
        if hafta_col < len(satir):
            h_val = satir[hafta_col]
            m = re.search(r"\d+", h_val)
            if m:
                aktif_hafta = int(m.group())
        if aktif_hafta is None:
            continue
        plan[aktif_hafta] = {
            "konu":    satir[konu_col]    if konu_col    is not None and konu_col < len(satir)    else "",
            "kazanim": satir[kazanim_col] if kazanim_col is not None and kazanim_col < len(satir) else "",
            "unite":   satir[unite_col]   if unite_col   is not None and unite_col < len(satir)   else "",
        }
    return plan


def _metin_yillik_coz(metin: str) -> dict:
    """Düz metinden hafta/konu/kazanım çıkarır."""
    plan = {}
    aktif_hafta = None
    for satir in metin.split("\n"):
        satir = satir.strip()
        m = re.search(r"(\d+)\.\s*hafta", satir, re.IGNORECASE)
        if m:
            aktif_hafta = int(m.group(1))
        if aktif_hafta and (
            re.search(r"kazanım|kazanim|F\.\d", satir, re.IGNORECASE)
        ):
            if aktif_hafta not in plan:
                plan[aktif_hafta] = {"konu": "", "kazanim": satir, "unite": ""}
            else:
                plan[aktif_hafta]["kazanim"] = satir
        elif aktif_hafta and len(satir) > 5 and aktif_hafta not in plan:
            plan[aktif_hafta] = {"konu": satir, "kazanim": "", "unite": ""}
    return plan
