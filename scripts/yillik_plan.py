"""
Yıllık Plan Okuyucu
===================
Excel yıllık planından her sınıf için doğru hafta konusunu çeker.
Sınıf seviyesine göre doğru sayfayı bulur (6, 7, 8).
"""

import re
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

PHET_ESLESTIRME = {
    "Kuvvet":     "https://phet.colorado.edu/tr/simulations/forces-and-motion-basics",
    "Newton":     "https://phet.colorado.edu/tr/simulations/forces-and-motion-basics",
    "Elektrik":   "https://phet.colorado.edu/tr/simulations/circuit-construction-kit-dc",
    "Direnç":     "https://phet.colorado.edu/tr/simulations/circuit-construction-kit-dc",
    "Işık":       "https://phet.colorado.edu/tr/simulations/bending-light",
    "Ses":        "https://phet.colorado.edu/tr/simulations/wave-on-a-string",
    "Basınç":     "https://phet.colorado.edu/tr/simulations/fluid-pressure-and-flow",
    "Atom":       "https://phet.colorado.edu/tr/simulations/build-an-atom",
    "DNA":        "https://learn.genetics.utah.edu/content/basics/",
    "Hücre":      "https://www.cellsalive.com",
    "Mıknatıs":   "https://phet.colorado.edu/tr/simulations/magnets-and-electromagnets",
    "Güneş":      "https://phet.colorado.edu/tr/simulations/gravity-and-orbits",
    "Isı":        "https://phet.colorado.edu/tr/simulations/states-of-matter",
    "Üreme":      "https://www.eba.gov.tr",
    "Madde":      "https://phet.colorado.edu/tr/simulations/states-of-matter",
}

VARSAYILAN_KONULAR = {
    "6": [
        {"hafta": range(1,5),   "unite":"1. Ünite", "konu":"Vücudumuzdaki Sistemler",    "kazanim":"F.6.1.1.1 — Sindirim sisteminin yapısını açıklar."},
        {"hafta": range(5,9),   "unite":"2. Ünite", "konu":"Kuvvet ve Hareket",           "kazanim":"F.6.2.1.1 — Sürtünme kuvvetini açıklar."},
        {"hafta": range(9,13),  "unite":"3. Ünite", "konu":"Madde ve Isı",                "kazanim":"F.6.3.1.1 — Isı ile sıcaklık farkını kavrar."},
        {"hafta": range(13,17), "unite":"4. Ünite", "konu":"Işık ve Ses",                 "kazanim":"F.6.4.1.1 — Işığın yansıma yasalarını açıklar."},
        {"hafta": range(17,21), "unite":"5. Ünite", "konu":"Elektrik",                    "kazanim":"F.6.5.1.1 — Statik elektriği keşfeder."},
        {"hafta": range(21,36), "unite":"6. Ünite", "konu":"Canlılar Dünyası",            "kazanim":"F.6.6.1.1 — Hücrenin yapısını açıklar."},
    ],
    "7": [
        {"hafta": range(1,5),   "unite":"1. Ünite", "konu":"Hücre Bölünmeleri",           "kazanim":"F.7.1.1.1 — Mitoz bölünmenin evrelerini açıklar."},
        {"hafta": range(5,9),   "unite":"2. Ünite", "konu":"Kuvvet ve Enerji",            "kazanim":"F.7.2.1.1 — Newton yasalarını açıklar."},
        {"hafta": range(9,13),  "unite":"3. Ünite", "konu":"Saf Madde ve Karışım",        "kazanim":"F.7.3.1.1 — Element ve bileşik farkını açıklar."},
        {"hafta": range(13,17), "unite":"4. Ünite", "konu":"Işığın Etkileşimi",           "kazanim":"F.7.4.1.1 — Işığın madde etkileşimini açıklar."},
        {"hafta": range(17,22), "unite":"5. Ünite", "konu":"Elektrik Devreleri",          "kazanim":"F.7.5.1.1 — Ohm yasasını açıklar."},
        {"hafta": range(22,36), "unite":"6. Ünite", "konu":"Canlılarda Üreme",            "kazanim":"F.7.6.1.1 — Üreme çeşitlerini karşılaştırır."},
    ],
    "8": [
        {"hafta": range(1,5),   "unite":"1. Ünite", "konu":"Mevsimler ve İklim",          "kazanim":"F.8.1.1.1 — Mevsimlerin oluşumunu açıklar."},
        {"hafta": range(5,9),   "unite":"2. Ünite", "konu":"DNA ve Genetik",              "kazanim":"F.8.2.1.1 — DNA yapısını açıklar."},
        {"hafta": range(9,13),  "unite":"3. Ünite", "konu":"Periyodik Sistem",            "kazanim":"F.8.3.1.1 — Periyodik sistemin düzenini açıklar."},
        {"hafta": range(13,17), "unite":"4. Ünite", "konu":"Basınç",                      "kazanim":"F.8.4.1.1 — Basınç kavramını açıklar."},
        {"hafta": range(17,22), "unite":"5. Ünite", "konu":"Ses",                         "kazanim":"F.8.5.1.1 — Ses dalgalarını açıklar."},
        {"hafta": range(22,36), "unite":"6. Ünite", "konu":"Manyetizma",                  "kazanim":"F.8.6.1.1 — Manyetik alanı açıklar."},
    ],
}


def simulasyon_bul(konu: str) -> str:
    for anahtar, link in PHET_ESLESTIRME.items():
        if anahtar.lower() in konu.lower():
            return link
    return "https://phet.colorado.edu/tr/simulations"


def okul_haftasi_hesapla(tarih: datetime) -> int:
    """Eylül başından itibaren kaçıncı hafta olduğunu hesaplar."""
    yil = tarih.year if tarih.month >= 9 else tarih.year - 1
    eylul1 = datetime(yil, 9, 1)
    # Eylül'ün ilk Pazartesi'si
    gun_farki = (7 - eylul1.weekday()) % 7
    if gun_farki == 0:
        gun_farki = 7
    ilk_pazartesi = datetime(yil, 9, min(gun_farki, 7))
    fark = (tarih - ilk_pazartesi).days
    return max(1, fark // 7 + 1)


def hafta_no_cikart(metin: str) -> int:
    m = re.search(r'(\d+)\.\s*[Hh]afta', str(metin))
    return int(m.group(1)) if m else 0


def excel_den_oku(dosya: Path, sinif: str, hafta_no: int) -> dict | None:
    """Excel yıllık planından belirli sınıf ve haftanın bilgisini çeker."""
    if not EXCEL_OK or not dosya.exists():
        return None

    seviye = sinif.split("-")[0]

    try:
        wb = openpyxl.load_workbook(dosya, data_only=True)
    except Exception:
        return None

    # Sınıf seviyesine göre sayfa bul
    hedef_sayfa = None
    for sayfa in wb.sheetnames:
        # "FEN BİLİMLERİ 6", "FEN BİLİMLERİ 7 (TYMM)" vb.
        if re.search(rf'\b{seviye}\b', sayfa):
            hedef_sayfa = sayfa
            break

    if not hedef_sayfa:
        return None

    ws = wb[hedef_sayfa]

    for satir in ws.iter_rows(values_only=True):
        hafta_val = satir[1] if len(satir) > 1 else None
        if not hafta_val:
            continue
        h = hafta_no_cikart(str(hafta_val))
        if h == hafta_no:
            unite   = str(satir[3]).strip() if satir[3] else ""
            konu    = str(satir[4]).strip() if satir[4] else ""
            kazanim = str(satir[5]).strip() if satir[5] else ""
            # None veya boş değerleri temizle
            if unite == "None": unite = ""
            if konu  == "None": konu  = ""
            if kazanim == "None": kazanim = ""
            if konu:
                return {
                    "unite":    unite[:60],
                    "konu":     konu[:60],
                    "kazanim":  kazanim[:120],
                }

    return None


class YillikPlanOkuyucu:
    def __init__(self, klasor: Path):
        self.klasor = klasor
        self._excel_cache = {}

    def _excel_dosyasi_bul(self) -> Path | None:
        """Klasördeki ilk Excel yıllık planını bulur."""
        if not self.klasor.exists():
            return None
        for f in self.klasor.rglob("*.xlsx"):
            if not f.name.startswith("~$"):
                return f
        return None

    def hafta_bilgisi_al(self, sinif: str, model: str, tarih: datetime) -> dict:
        seviye = sinif.split("-")[0]
        hafta_no = okul_haftasi_hesapla(tarih)

        # 1. Excel dosyasından oku
        excel = self._excel_dosyasi_bul()
        if excel:
            sonuc = excel_den_oku(excel, sinif, hafta_no)
            if sonuc and sonuc["konu"]:
                print(f"   ✓  [{sinif}] Excel'den okundu: {sonuc['konu'][:40]}")
                return {
                    "unite":      sonuc["unite"],
                    "konu":       sonuc["konu"],
                    "kazanim":    sonuc["kazanim"],
                    "simulasyon": simulasyon_bul(sonuc["konu"]),
                    "model":      model,
                    "hafta":      hafta_no,
                }

        # 2. Fallback: yerleşik plan
        plan = VARSAYILAN_KONULAR.get(seviye, [])
        for giris in plan:
            if hafta_no in giris["hafta"]:
                konu_idx = (hafta_no - giris["hafta"].start) % 1
                konu = giris["konu"]
                return {
                    "unite":      giris["unite"],
                    "konu":       konu,
                    "kazanim":    giris["kazanim"],
                    "simulasyon": simulasyon_bul(konu),
                    "model":      model,
                    "hafta":      hafta_no,
                }

        return {
            "unite": "—", "konu": "Genel Tekrar",
            "kazanim": "Kazanım yıllık plandan alınacak",
            "simulasyon": "", "model": model, "hafta": hafta_no,
        }

